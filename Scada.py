# необходимые библиотеки

from tkinter import *
import random
from openpyxl import *
from PIL import*
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from datetime import datetime, timedelta
from tkinter import ttk

# начальные значения по умолчанию
ArrX = [0.0]
SensorTemp_RTD_array = [20.0]
SensorTemp_Cuprum_array = [20.0]
SensorTemp_TPL_array = [20.0]
SensorTemp_TPK_array = [20.0]
SensorTemp_RT100_array = [20.0]
ValueTemp = []

CountBtnClck = 1
CountBtnClck1 = 1
CountBtnClck2 = 1
CountBtnClck3 = 1
id = None
temp_id = None

db = Workbook()
ActiveList = db.active

# сохранение в excel
def save_data_to_excel():
    global ArrX, SensorTemp_RTD_array, SensorTemp_Cuprum_array, SensorTemp_TPL_array, SensorTemp_TPK_array, SensorTemp_RT100_array, db, ActiveList
    ActiveList.delete_rows(1, ActiveList.max_row)
    ActiveList['A1'] = 'Время'
    ActiveList['B1'] = 'RTD °С'
    ActiveList['C1'] = 'Cuprum °С'
    ActiveList['D1'] = 'TPL °С'
    ActiveList['E1'] = 'TPK °С'
    ActiveList['F1'] = 'RT100 °С'
    for i in range(len(ArrX)):
        ActiveList.cell(row=i+2, column=1).value = ArrX[i]
        ActiveList.cell(row=i+2, column=2).value = SensorTemp_RTD_array[i]
        ActiveList.cell(row=i+2, column=3).value = SensorTemp_Cuprum_array[i]
        ActiveList.cell(row=i+2, column=4).value = SensorTemp_TPL_array[i]
        ActiveList.cell(row=i+2, column=5).value = SensorTemp_TPK_array[i]
        ActiveList.cell(row=i+2, column=6).value = SensorTemp_RT100_array[i]
    db.save('Database.xlsx')
    LabelData.configure(text = 'Данные записаны')

# функция съема значения с ползунка
def scaleget(newVal):
    scale_tittle.configure(text=f'Нагрев:{newVal}%', font=('Consolas', 10))
    if int(newVal) > 0:
        scale_tittle.configure(bg='green')
    else:
        scale_tittle.configure(bg='grey80')
    for i in range(9, -1, -1):
        if i < int(newVal) // 10:
            List[i].configure(bg='green')
        else:
            List[i].configure(bg='red')
# текущее состояние кнопок
def BtnChangeState(text):
    global CountBtnClck, id, temp_id
    if text == '1':
        CountBtnClck += 1
        if CountBtnClck % 2 == 0:
            AutoRegulation['bg'] = 'green'
        else:
            AutoRegulation['bg'] = 'grey'
    if text == '3':
        global CountBtnClck2
        CountBtnClck2 += 1
        if CountBtnClck2 % 2 == 0:
            btn['bg'] = 'green'
            create_plot()
            update_plot()
            if SensorTemp_RT100_array[-1] > 60:
                frame.after_cancel(temp_id)
        else:
            btn['bg'] = 'grey'
            frame.after_cancel(id)
            destroy_plot()
    if text == '4':
        global CountBtnClck3
        CountBtnClck3 += 1
        if CountBtnClck3 % 2 == 0:
            RedrawBtn['bg'] = 'green'
        else:
            RedrawBtn['bg'] = 'grey'

# создание графика
def create_plot():
    global fig, plot, canvas, start_time
    fig = Figure(figsize=(7,4), dpi=100)
    plot = fig.add_subplot(1, 1, 1)
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas.get_tk_widget().place(x=150, y=150)
    start_time = datetime.now()
# обновление графика
def update_plot():
    global ArrX, SensorTemp_RTD_array, SensorTemp_Cuprum_array, SensorTemp_TPL_array, SensorTemp_TPK_array, SensorTemp_RT100_array, CountBtnClck2, id, start_time, table, temp_id
    elapsed_time = datetime.now() - start_time
    x = elapsed_time.total_seconds()
    if CountBtnClck % 2 == 1:
        SensorTemp_RTD_array.append(scale.get()*0.01+SensorTemp_RTD_array[-1])
        SensorTemp_Cuprum_array.append(scale.get()*0.02+SensorTemp_Cuprum_array[-1])
        SensorTemp_TPL_array.append(scale.get()*0.03+SensorTemp_TPL_array[-1])
        SensorTemp_TPK_array.append(scale.get()*0.04+SensorTemp_TPK_array[-1])
        SensorTemp_RT100_array.append(scale.get()*0.05+SensorTemp_RT100_array[-1])
        
    if CountBtnClck % 2 == 0:
        if SensorTemp_RTD_array[-1] <= 15:
            SensorTemp_RTD_array.append(15)
        else:
            SensorTemp_RTD_array.append(SensorTemp_RTD_array[-1] - random.randint(0,1))
        if SensorTemp_Cuprum_array[-1] <= 15:
            SensorTemp_Cuprum_array.append(15)
        else:
            SensorTemp_Cuprum_array.append(SensorTemp_Cuprum_array[-1] - random.randint(1,3))
        if SensorTemp_TPL_array[-1] <= 15:
            SensorTemp_TPL_array.append(15)
        else:
            SensorTemp_TPL_array.append(SensorTemp_TPL_array[-1] - random.randint(1,2))
        if SensorTemp_TPK_array[-1] <= 15:
            SensorTemp_TPK_array.append(15)
        else:
            SensorTemp_TPK_array.append(SensorTemp_TPK_array[-1] - random.randint(1,4))
        if SensorTemp_RT100_array[-1] <= 15:
            SensorTemp_RT100_array.append(15)
        else:
            SensorTemp_RT100_array.append(SensorTemp_RT100_array[-1] - random.randint(1,2))
        
    state = 'normal'
    if SensorTemp_RT100_array[-1] > 60:
        state = 'Immediately stop heating!'
    else:
        state = 'normal'
    ValueTemp.append([SensorTemp_RT100_array[-1], x, state])
    for i in ValueTemp:
        table.insert("", END, values=i)
        ValueTemp.clear()
    ArrX.append(x)
    plot.clear()
    plot.set_xlabel('Время [cек]')
    plot.set_ylabel('Температура [°C]')
    plot.plot(ArrX, SensorTemp_RTD_array, 'o-c',
            ArrX, SensorTemp_Cuprum_array, 'o--b',
            ArrX, SensorTemp_TPL_array, 'o-.r',
            ArrX, SensorTemp_TPK_array, 'o:g',
            ArrX, SensorTemp_RT100_array, 'o:k')
    canvas.draw()
    labelRTD.config(text = f"RTD: {round(SensorTemp_RTD_array[-1], 2)}°C")
    labelCuprum.config(text = f"Cuprum: {round(SensorTemp_Cuprum_array[-1],2 )}°C")
    labelTPL.config(text = f"TPL: {round(SensorTemp_TPL_array[-1], 2)}°C")
    labelTPK.config(text = f"TPK: {round(SensorTemp_TPK_array[-1], 2)}°C")
    labelRT100.config(text = f"RT100: {round(SensorTemp_RT100_array[-1], 2)}°C")
    if SensorTemp_RT100_array[-1] < 65:
        id = frame.after(1000, update_plot)
    else:
        frame.after_cancel(id)

# сброс графика
def reset_plot():
    global ArrX, SensorTemp_RTD_array, SensorTemp_Cuprum_array, SensorTemp_TPL_array, SensorTemp_TPK_array, SensorTemp_RT100_array, start_time
    start_time = datetime.now()
    ArrX = [0.0]
    SensorTemp_RTD_array = [20.0]
    SensorTemp_Cuprum_array = [20.0]
    SensorTemp_TPL_array = [20.0]
    SensorTemp_TPK_array = [20.0]
    SensorTemp_RT100_array = [20.0]
    plot.clear()
    canvas.draw()

# Скрытие графика
def destroy_plot():
    global ArrX, SensorTemp_RTD_array, SensorTemp_Cuprum_array, SensorTemp_TPL_array, SensorTemp_TPK_array, SensorTemp_RT100_array, start_time, fig, plot
    start_time = datetime.now()
    ArrX = [0.0]
    SensorTemp_RTD_array = [20.0]
    SensorTemp_Cuprum_array = [20.0]
    SensorTemp_TPL_array = [20.0]
    SensorTemp_TPK_array = [20.0]
    SensorTemp_RT100_array = [20.0]
    fig = None
    canvas.get_tk_widget().destroy()

# главное окно
main = Tk()
main.state('zoomed')
main.title("SCADA система температурных датчиков")
window_width = 800
window_height = 600
main.geometry('%dx%d' % (window_width, window_height))

# контейнер с мнемосхемой
photo = PhotoImage(file='D:\Учеба\Программное обеспечение АС/new.png')
frame = Frame(main, width=1920, height=1080, bg='MistyRose1')
label = Label(frame, image=photo)
label.place(x=0, y=0)
frame.pack()

# описание кнопок
AutoRegulation = Button(frame, text='Вкл. вентилирование', bg = 'grey', width=22, height=6, state = None, command=lambda: BtnChangeState('1'))
btn = Button(frame, text='Запуск/стоп', bg = 'grey', width=21, height=2, state = None, command=lambda: BtnChangeState('3'))
RedrawBtn = Button(frame, text='Сбросить значения', bg = 'grey', width=21, height=2, command=reset_plot)
SaveData = Button(frame, text= 'Запись значений в excel', bg = 'blue', width= 20, height=2, command=save_data_to_excel)
SaveData.place(x= 350, y=575)
AutoRegulation.place(x=1005, y=610)
btn.place(x=50, y=50)
RedrawBtn.place(x = 150, y = 575)

# описание ползунка
scale_tittle = Label(frame, text = 'Нагрев %', width=10, height=1, bg = 'grey80')
scale_tittle.place(x = 1057, y = 348)
scale = Scale(frame, orient='vertical', from_= 100, to=0, width=70, length=220, showvalue=0, sliderlength=20, sliderrelief='raised', command=scaleget)
scale.place(x=1057, y=368)

# лейблы для датчиков

labelRTD = Label(frame, text = "", width = 11, height = 1, bg = 'grey80')
labelCuprum = Label(frame, text = "", width = 15, height = 1, bg = 'grey80')
labelTPL = Label(frame, text = "", width = 13, height = 1, bg = 'grey80')
labelTPK = Label(frame, text = "", width = 13, height = 1, bg = 'grey80')
labelRT100 = Label(frame, text = "", width = 15, height = 1, bg = 'grey80')
LabelData = Label(frame, text='', width=50, height=5, bg='grey80', font='Arial')

labelRTD.place(x=1063, y =304)
labelCuprum.place(x=1055, y=246)
labelTPL.place(x=1241,y=217)
labelTPK.place(x=1305, y=323)
labelRT100.place(x=1178, y=333)
LabelData.place(x=150, y= 700)

# создание таблицы с аварийными значениями
columns = ('value', 'time', 'comments')
table = ttk.Treeview(columns=columns, show='headings')
table.heading('value', text = 'Значение °С')
table.heading('time', text = 'Время в секундах')
table.heading('comments', text = 'Состояние')
table.place(x=1000, y=750)


# лейблы и фрейм для нагревательного элемента
frame_temp = Frame(frame, width=220, height=226)
frame_temp.place(x=1155, y=363)
currenty = 0
List = []
for i in range(10):
    lab = Label(frame_temp, width=24, height=1, bg='red', borderwidth=1, highlightcolor='black', highlightbackground='black', highlightthickness=2)
    lab.pack(side='bottom')
    currenty += 22
    List.append(lab)

main.mainloop()